from openpyxl import load_workbook
wb = load_workbook(filename = 'test.xlsx')
ws = wb.active
ws['B4'].value = "Hello"

print(ws['A1'].value)